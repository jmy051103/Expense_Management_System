# expenses/views.py
# expenses/views.py (top)
from decimal import Decimal, InvalidOperation

import datetime
import io

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.core.exceptions import PermissionDenied
from django.core.files.storage import default_storage
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import F
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils.http import url_has_allowed_host_and_scheme
from django.views.decorators.http import require_POST

from PIL import Image as PILImage
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side

from .forms import ExpenseReportForm, ExpenseItemFormSet, ContractForm
from .models import ExpenseReport, Contract, ContractImage, ContractItem


def _open_pil_from_field(file_field):
    if not file_field:
        return None
    # 1) 파일 경로로 열기
    path = getattr(file_field, "path", None)
    try:
        if path:
            return PILImage.open(path)
    except Exception:
        pass
    # 2) 스토리지에서 바이너리로 열기
    try:
        name = getattr(file_field, "name", None)
        if not name:
            return None
        with default_storage.open(name, "rb") as f:
            bio = io.BytesIO(f.read())
            bio.seek(0)
            return PILImage.open(bio)
    except Exception:
        return None

def _is_approver(user) -> bool:
    """approver 그룹 또는 superuser라면 True"""
    return user.is_superuser or user.groups.filter(name="approver").exists()

# ------------------- 기존 보고서(ExpenseReport) 뷰들 그대로 -------------------
@login_required
def report_list(request):
    reports = (
        ExpenseReport.objects
        .select_related("creator")
        .prefetch_related("items")
        .order_by("-created_at")
    )
    can_delete = _is_approver(request.user)
    return render(request, "expenses/report_list.html", {
        "reports": reports,
        "can_delete": can_delete,
    })

@login_required
def report_create(request):
    if request.method == "POST":
        form = ExpenseReportForm(request.POST)
        formset = ExpenseItemFormSet(request.POST)
        if form.is_valid() and formset.is_valid():
            with transaction.atomic():
                report = form.save(commit=False)
                report.creator = request.user
                report.save()
                formset.instance = report
                formset.save()
            return redirect(reverse("expenses:report_detail", args=[report.id]))
    else:
        form = ExpenseReportForm()
        formset = ExpenseItemFormSet()

    return render(
        request,
        "expenses/report_form.html",
        {"form": form, "formset": formset, "is_edit": False},
    )

@login_required
def report_edit(request, pk):
    report = get_object_or_404(ExpenseReport, pk=pk)
    if request.method == "POST":
        form = ExpenseReportForm(request.POST, instance=report)
        formset = ExpenseItemFormSet(request.POST, instance=report)
        if form.is_valid() and formset.is_valid():
            with transaction.atomic():
                form.save()
                formset.save()
            return redirect(reverse("expenses:report_detail", args=[report.id]))
    else:
        form = ExpenseReportForm(instance=report)
        formset = ExpenseItemFormSet(instance=report)

    return render(
        request,
        "expenses/report_form.html",
        {"form": form, "formset": formset, "is_edit": True, "report": report},
    )

@login_required
def report_detail(request, pk):
    report = get_object_or_404(
        ExpenseReport.objects.select_related("creator").prefetch_related("items"),
        pk=pk,
    )
    can_delete = _is_approver(request.user)
    return render(request, "expenses/report_detail.html", {
        "report": report,
        "can_delete": can_delete,
    })

@login_required
@require_POST
def report_delete(request, pk):
    if not _is_approver(request.user):
        raise PermissionDenied("You do not have permission to delete this report.")
    report = get_object_or_404(ExpenseReport, pk=pk)
    report.delete()
    return redirect("expenses:report_list")

# ------------------- 여기부터 계약(Contract) 뷰 -------------------

@login_required
def add_contract(request):
    """계약정보 등록 + 이미지 다중 업로드.
       '저장하기' => draft, '품의요청' => submitted
    """
    sales_people = (
        User.objects.filter(is_active=True)
        .select_related("profile")
        .order_by("first_name", "username")
    )

    if request.method == "POST":
        is_submit = request.POST.get("submit_final") == "1"
        status = "submitted" if is_submit else "draft"

        form = ContractForm(request.POST)
        if form.is_valid():
            with transaction.atomic():
              
                contract = form.save(commit=False)
                contract.writer = request.user
                contract.status = status
                contract.title = contract.customer_company or "무제 계약"
                contract.save()

                for f in request.FILES.getlist("images"):
                    ContractImage.objects.create(contract=contract, original=f)
                
                names = request.POST.getlist("item_name[]") or []
                qtys  = request.POST.getlist("qty[]") or []
                specs = request.POST.getlist("spec[]") or []
                su    = request.POST.getlist("sell_unit[]") or []
                st    = request.POST.getlist("sell_total[]") or []
                bu    = request.POST.getlist("buy_unit[]") or []
                bt    = request.POST.getlist("buy_total[]") or []
                vend  = request.POST.getlist("vendor[]") or []
                vat   = request.POST.getlist("item_vat_mode[]") or []

                def get(lst, i, default=""):
                    return lst[i] if i < len(lst) else default
                
                for i in range(len(names)):
                    name = (get(names, i, "") or "").strip()
                    qty  = _i(get(qtys, i, 0))
                    if not name or qty <= 0:
                        continue  # 서버에서도 최소 검증

                    sell_unit  = _d(get(su, i, 0))
                    buy_unit   = _d(get(bu, i, 0))
                    sell_total = _d(get(st, i, 0))
                    buy_total  = _d(get(bt, i, 0))

                    # 서버에서 금액 재계산(신뢰성 향상): 단가/수량이 있으면 금액 덮어쓰기
                    if qty and sell_unit:
                        sell_total = (sell_unit * qty).quantize(Decimal("1."))  # 반올림 규칙은 필요시 변경
                    if qty and buy_unit:
                        buy_total  = (buy_unit * qty).quantize(Decimal("1."))

                    ContractItem.objects.create(
                        contract=contract,
                        name=name,
                        qty=qty,
                        spec=get(specs, i, "") or "",
                        sell_unit=sell_unit,
                        sell_total=sell_total,
                        buy_unit=buy_unit,
                        buy_total=buy_total,
                        vendor=get(vend, i, "") or "",
                        vat_mode=(get(vat, i, "separate") or "separate"),
                    )

            return redirect("expenses:contract_detail", pk=contract.pk)
        else:
            # 폼 에러를 템플릿에서 표시할 수 있게 넘김
            ctx = {
                "sales_people": sales_people,
                "customer_managers": [],
                "form_errors": form.errors,
            }
            return render(request, "add_contract.html", ctx)

    # GET
    ctx = {"sales_people": sales_people, "customer_managers": []}
    return render(request, "add_contract.html", ctx)

@login_required
def contract_detail(request, pk):
    """계약 상세"""
    contract = get_object_or_404(
        Contract.objects
        .select_related("writer", "sales_owner")
        .prefetch_related("images", "items"),  # ← items도 미리
        pk=pk
    )
    return render(request, "contract_detail.html", {"contract": contract})

@login_required
def contract_edit(request, pk):
    contract = get_object_or_404(
        Contract.objects
        .select_related("writer", "sales_owner")
        .prefetch_related("images", "items"),
        pk=pk
    )

    # next 계산(생략 가능: 기존 그대로 사용)
    if request.method == "POST":
        raw_next = request.POST.get("next") or request.GET.get("next")
    else:
        raw_next = (
            request.GET.get("next")
            or request.META.get("HTTP_REFERER")
            or reverse("expenses:contract_detail", args=[contract.pk])
        )
    next_url = raw_next
    if not url_has_allowed_host_and_scheme(
        next_url, allowed_hosts={request.get_host()}, require_https=request.is_secure()
    ):
        next_url = reverse("expenses:contract_detail", args=[contract.pk])

    # 권한
    if not (request.user.is_superuser or request.user == contract.writer):
        messages.error(request, "수정 권한이 없습니다. (작성자만 수정 가능)")
        return redirect(next_url)

    sales_people = (
        User.objects.filter(is_active=True)
        .select_related("profile")
        .order_by("first_name", "username")
    )

    if request.method == "POST":
        is_submit = request.POST.get("submit_final") == "1"
        form = ContractForm(request.POST, instance=contract)
        if form.is_valid():
            with transaction.atomic():
                # 기존 상태 보관
                prev_status = contract.status

                # 본문 저장
                contract = form.save(commit=False)
                contract.title = contract.customer_company or (contract.title or "무제 계약")

                # ✅ 상태 결정 로직 변경
                if is_submit:
                    # 사용자가 '결재요청' 버튼을 누른 경우에만 제출 상태로
                    contract.status = "submitted"
                else:
                    # 그 외(일반 저장)는 기존 상태를 그대로 유지
                    contract.status = prev_status

                contract.save()

                # 이미지 삭제/추가
                del_ids = request.POST.getlist("del_image_ids[]")
                if del_ids:
                    ContractImage.objects.filter(contract=contract, id__in=del_ids).delete()
                for f in request.FILES.getlist("images"):
                    ContractImage.objects.create(contract=contract, original=f)

                # 품목 재작성(서버 재계산)
                contract.items.all().delete()
                names = request.POST.getlist("item_name[]") or []
                qtys  = request.POST.getlist("qty[]") or []
                specs = request.POST.getlist("spec[]") or []
                su    = request.POST.getlist("sell_unit[]") or []
                st    = request.POST.getlist("sell_total[]") or []
                bu    = request.POST.getlist("buy_unit[]") or []
                bt    = request.POST.getlist("buy_total[]") or []
                vend  = request.POST.getlist("vendor[]") or []
                vat   = request.POST.getlist("item_vat_mode[]") or []

                def get(lst, i, default=""):
                    return lst[i] if i < len(lst) else default

                for i in range(len(names)):
                    name = (get(names, i, "") or "").strip()
                    qty  = _i(get(qtys, i, 0))
                    if not name or qty <= 0:
                        continue

                    sell_unit  = _d(get(su, i, 0))
                    buy_unit   = _d(get(bu, i, 0))
                    sell_total = _d(get(st, i, 0))
                    buy_total  = _d(get(bt, i, 0))

                    if qty and sell_unit:
                        sell_total = (sell_unit * qty).quantize(Decimal("1."))
                    if qty and buy_unit:
                        buy_total  = (buy_unit  * qty).quantize(Decimal("1."))

                    ContractItem.objects.create(
                        contract=contract,
                        name=name,
                        qty=qty,
                        spec=get(specs, i, "") or "",
                        sell_unit=sell_unit,
                        sell_total=sell_total,
                        buy_unit=buy_unit,
                        buy_total=buy_total,
                        vendor=get(vend, i, "") or "",
                        vat_mode=(get(vat, i, "separate") or "separate"),
                    )

            messages.success(request, "계약이 저장되었습니다.")
            return redirect(next_url)

        # 폼 에러
        ctx = {
            "sales_people": sales_people,
            "customer_managers": [],
            "form_errors": form.errors,
            "contract": contract,
            "is_edit": True,
            "next": next_url,
        }
        return render(request, "add_contract.html", ctx)

    # GET
    ctx = {
        "sales_people": sales_people,
        "customer_managers": [],
        "contract": contract,
        "is_edit": True,
        "next": next_url,
    }
    return render(request, "add_contract.html", ctx)

@login_required
@require_POST
def contract_delete(request, pk):
    """계약 삭제: 작성자 또는 superuser만. 삭제 후 원래 보던 목록으로 리다이렉트"""
    # 사용자가 보던 페이지(쿼리스트링 포함)를 next로 전달받거나, 없으면 Referer 사용
    next_url = (
        request.POST.get("next")
        or request.GET.get("next")
        or request.META.get("HTTP_REFERER")
        or reverse("expenses:contract_list")
    )
    # 안전한 내부 URL인지 검증
    if not url_has_allowed_host_and_scheme(
        next_url,
        allowed_hosts={request.get_host()},
        require_https=request.is_secure(),
    ):
        next_url = reverse("expenses:contract_list")

    contract = get_object_or_404(Contract, pk=pk)

    if not (request.user.is_superuser or request.user == contract.writer):
        from django.contrib import messages
        messages.error(request, "삭제 권한이 없습니다. (작성자만 삭제 가능)")
        return redirect(next_url)

    contract.delete()
    from django.contrib import messages
    messages.success(request, "계약이 성공적으로 삭제되었습니다.")
    return redirect(next_url)

@login_required
def contract_list(request):
    """계약 목록 (검색 + per_page 선택 + 숫자 페이지네이션 + 행 선택 체크박스만)"""
    sales_people = (
        User.objects.filter(is_active=True)
        .select_related("profile")
        .order_by("first_name", "username")
    )

    # ✅ 기본 쿼리셋은 먼저 만들기
    qs = (
        Contract.objects
        .select_related("writer", "sales_owner")
        .prefetch_related("images", "items")
        .order_by(
            F("collect_invoice_date").desc(nulls_last=True),
            "-created_at",
            "-id",
        )
    )

    # ===== 검색/필터 =====
    date_from   = (request.GET.get("date_from") or "").strip()
    date_to     = (request.GET.get("date_to") or "").strip()
    q_customer  = (request.GET.get("q_customer") or "").strip()
    q_vendor    = (request.GET.get("q_vendor") or "").strip()
    owner_id    = (request.GET.get("owner") or "").strip()  
    q_item      = (request.GET.get("q_item") or "").strip()
    contract_no = (request.GET.get("contract_no") or "").strip()
    status      = (request.GET.get("status") or "").strip()

    if date_from:
        qs = qs.filter(created_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(created_at__date__lte=date_to)
    if q_customer:
        qs = qs.filter(customer_company__icontains=q_customer)
    if q_vendor:
        qs = qs.filter(items__vendor__icontains=q_vendor).distinct()
    if owner_id:
        # ✅ 작성자(id)로 필터
        try:
            qs = qs.filter(writer_id=int(owner_id))
        except (TypeError, ValueError):
            pass
    if q_item:
        qs = qs.filter(items__name__icontains=q_item).distinct()
    if contract_no:
        qs = qs.filter(contract_no__icontains=contract_no)
    if status:
        qs = qs.filter(status=status)

    # ===== per_page & 페이지네이션 =====
    per_page_options = [10, 20, 30, 50, 100]
    try:
        per_page = int(request.GET.get("per_page", 10))
    except (TypeError, ValueError):
        per_page = 10
    if per_page not in per_page_options:
        per_page = 10

    paginator = Paginator(qs, per_page)
    page_number = request.GET.get("page") or 1
    page_obj = paginator.get_page(page_number)

    # page 파라미터 제거한 쿼리스트링 (숫자 페이지네이션 링크에 사용)
    qs_keep = request.GET.copy()
    qs_keep.pop("page", None)
    qs_without_page = qs_keep.urlencode()

    return render(request, "contract_list.html", {
        "page_obj": page_obj,
        "contracts": page_obj,
        "per_page_options": per_page_options,
        "per_page": per_page,
        "qs": qs_without_page,
        "sales_people": sales_people,
    })

def _d(v):
    """'1,234' -> Decimal('1234'); blanks -> Decimal('0')"""
    s = (str(v) if v is not None else "").replace(",", "").strip()
    try:
        return Decimal(s) if s else Decimal("0")
    except InvalidOperation:
        return Decimal("0")
    
def _i(v):
    """int parser that tolerates commas/blank"""
    try:
        return int(_d(v))
    except Exception:
        return 0
    

@login_required
def contract_export(request):
    """
    계약 목록 엑셀 다운로드.
    - 체크된 행이 있으면 ?ids=1,2,3 만 내보냄
    - 없으면 현재 검색필터가 적용된 전체를 내보냄
    - 이익금액/이익율은 품목 합계로 즉석 계산
    - 사진: 여러 장이면 모두 삽입(세로 병합된 '사진' 칸에 세로로 쌓음)
    - 한 계약에 품목이 여러 개면: 계약 공통 칼럼은 세로 병합, 품목/수량 등만 행별 기재
    """
    from decimal import Decimal
    import io, datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
    from openpyxl.drawing.image import Image as XLImage
    from django.http import HttpResponse
    from django.db.models import F

    # 기본 쿼리 (contract_list와 동일 정렬)
    qs = (
        Contract.objects
        .select_related("writer", "sales_owner")
        .prefetch_related("images", "items")
        .order_by(
            F("collect_invoice_date").desc(nulls_last=True),
            "-created_at",
            "-id",
        )
    )

    # ===== 선택된 id 우선 처리 =====
    raw_ids = []
    raw_ids += request.GET.getlist("ids")
    ids_csv = request.GET.get("ids", "")
    if ids_csv:
        raw_ids += ids_csv.split(",")
    ids = []
    for t in raw_ids:
        for piece in str(t).split(","):
            piece = piece.strip()
            if piece.isdigit():
                ids.append(int(piece))
    if ids:
        qs = qs.filter(id__in=ids)

    # ===== 검색/필터 =====
    date_from   = (request.GET.get("date_from") or "").strip()
    date_to     = (request.GET.get("date_to") or "").strip()
    q_customer  = (request.GET.get("q_customer") or "").strip()
    q_vendor    = (request.GET.get("q_vendor") or "").strip()
    owner_id    = (request.GET.get("owner") or "").strip()
    q_item      = (request.GET.get("q_item") or "").strip()
    contract_no = (request.GET.get("contract_no") or "").strip()
    status      = (request.GET.get("status") or "").strip()

    if date_from:
        qs = qs.filter(created_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(created_at__date__lte=date_to)
    if q_customer:
        qs = qs.filter(customer_company__icontains=q_customer)
    if q_vendor:
        qs = qs.filter(items__vendor__icontains=q_vendor).distinct()
    if owner_id:
        qs = qs.filter(sales_owner_id=owner_id)
    if q_item:
        qs = qs.filter(items__name__icontains=q_item).distinct()
    if contract_no:
        qs = qs.filter(contract_no__icontains=contract_no)
    if status:
        qs = qs.filter(status=status)

    # ===== 워크북/워크시트 =====
    wb = Workbook()
    ws = wb.active
    ws.title = "계약목록"

    header_fill = PatternFill("solid", fgColor="7EA0B8")
    line_color = "6F8EA6"
    thin = Side(style="thin", color=line_color)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 돈 서식
    existing_style_names = {getattr(s, "name", str(s)) for s in wb.named_styles}
    if "krw" not in existing_style_names:
        money = NamedStyle(name="krw")
        money.number_format = '#,##0"원"'
        try:
            wb.add_named_style(money)
        except ValueError:
            pass

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right  = Alignment(horizontal="right", vertical="center")

    # A~I: 계약 공통 / J~Q: 품목 / R: 사진
    headers = [
        "계약번호","상태","매출처","담당자",
        "작성자","작성일","마감월",
        "이익금액","이익율",
        "품목","규격","수량",
        "매출단가","매출금액",
        "매입단가","매입금액",
        "매입처",
        "사진",
    ]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = header_fill
        c.alignment = center
        c.border = border
    ws.row_dimensions[1].height = 22

    widths = [13,10,18,12,12,11,10,13,9,28,16,9,13,14,13,14,16,12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    # PIL 로더
    def _open_pil_from_field(f):
        try:
            if not f:
                return None
            f.open("rb")
            from PIL import Image as PILImage
            return PILImage.open(f).copy()
        except Exception:
            return None
        finally:
            try:
                f.close()
            except Exception:
                pass

    def add_images_merged(top_row, bottom_row, col_letter, pil_list):
        """세로 병합된 셀에 이미지를 여러 장 세로로 배치"""
        if not pil_list:
            return
        # 행 높이 조금 넉넉히 잡기 (원본 그대로 쓰면 150~200px 권장)
        rows = list(range(top_row, bottom_row + 1))
        for r in rows:
            ws.row_dimensions[r].height = 120  # 행 높이 크게 조정

        for pil in pil_list:
            img = pil.convert("RGB")
            # 🔽 여기서 크기를 줄이지 않거나, 더 크게 지정
            img.thumbnail((200, 200))   # 필요에 맞게 조정
            bio = io.BytesIO()
            img.save(bio, format="PNG")
            bio.seek(0)
            xlimg = XLImage(bio)
            ws.add_image(xlimg, f"{col_letter}{top_row}")

    row = 2
    for c in qs:
        # 품목 합계로 이익/이익율 계산
        sell_sum = Decimal("0")
        buy_sum  = Decimal("0")
        items_all = list(c.items.all())
        for it in items_all:
            sell_sum += Decimal(it.sell_total or 0)
            buy_sum  += Decimal(it.buy_total or 0)
        profit = sell_sum - buy_sum
        margin_rate = (profit / sell_sum * Decimal("100")) if sell_sum > 0 else None

        # 이미지 전부 수집(thumb → medium → original 우선순위)
        pil_list = []
        if c.images.exists():
            for ci in c.images.all():
                for field in (getattr(ci, "thumb", None), getattr(ci, "medium", None), getattr(ci, "original", None)):
                    pil = _open_pil_from_field(field)
                    if pil:
                        pil_list.append(pil)
                        break

        # 품목이 없으면 빈 한 줄 보장
        items = items_all or [None]
        start_row = row
        for idx, it in enumerate(items):
            # 첫 행엔 계약 공통 칼럼, 이후 행은 비워두고 나중에 세로 병합
            if idx == 0:
                base_vals = [
                    (c.contract_no or c.id),
                    c.get_status_display(),
                    (c.customer_company or c.title or ""),
                    (c.customer_manager or ""),
                    (c.writer.first_name or c.writer.username) if c.writer_id else "",
                    c.created_at.strftime("%Y-%m-%d") if c.created_at else "",
                    c.margin_month or "",
                    int(profit) if profit is not None else "",
                    (f"{margin_rate:.2f}%") if margin_rate is not None else "",
                ]
                ws.append(
                    base_vals + [
                        it.name if it else "",
                        it.spec if it else "",
                        it.qty if it else "",
                        it.sell_unit if it else "",
                        it.sell_total if it else "",
                        it.buy_unit if it else "",
                        it.buy_total if it else "",
                        it.vendor if it else "",
                        "",  # 사진 (아래서 이미지 넣음)
                    ]
                )
            else:
                # 공통 칸은 빈 값(병합 예정), 품목 칸만 기록
                ws.append(
                    [""] * 9 + [
                        it.name if it else "",
                        it.spec if it else "",
                        it.qty if it else "",
                        it.sell_unit if it else "",
                        it.sell_total if it else "",
                        it.buy_unit if it else "",
                        it.buy_total if it else "",
                        it.vendor if it else "",
                        "",
                    ]
                )

            # 스타일/정렬/테두리/금액 서식
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                if col in (1,2,3,4,5,6,7,10,11,17):
                    cell.alignment = left
                elif col in (12,):
                    cell.alignment = center
                else:
                    cell.alignment = right
                if col in (13,14,15,16,8):  # 금액 계열 + 이익금액
                    v = cell.value
                    if isinstance(v, (int, float, Decimal)):
                        cell.style = "krw"
                        if isinstance(v, Decimal):
                            cell.value = float(v)
            row += 1

        end_row = row - 1

        # 세로 병합: A~I(공통) + R(사진)
        if end_row > start_row:
            for col_idx in list(range(1, 10)) + [18]:
                col_letter = ws.cell(row=1, column=col_idx).column_letter
                ws.merge_cells(f"{col_letter}{start_row}:{col_letter}{end_row}")
                # 병합 좌상단 정렬 유지
                ws.cell(row=start_row, column=col_idx).alignment = left if col_idx in (1,2,3,4,5,6,7,10,11,17) else right

        # 이미지 모두 삽입(병합된 사진 칼럼에)
        if pil_list:
            col_letter = ws.cell(row=1, column=18).column_letter  # 사진 열
            add_images_merged(start_row, end_row, col_letter, pil_list)

    today = datetime.date.today().strftime("%Y%m%d")
    fname = f"contract_export_{today}.xlsx"
    mem = io.BytesIO()
    wb.save(mem)
    mem.seek(0)
    return HttpResponse(
        mem.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )